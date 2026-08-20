"""
Генерация XLSX-таблицы из базы данных мониторинга.

Использование:
    python scripts/build_xlsx.py

Читает data/monitoring_data.json, создаёт output/monitoring_base.xlsx
"""

import json
import os
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_FILE = os.path.join(PROJECT_ROOT, "data", "monitoring_data.json")
OUTPUT_FILE = os.path.join(PROJECT_ROOT, "output", "monitoring_base.xlsx")


def load_data():
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def build_xlsx(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "База знаний"

    # Стили
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    wrap_align = Alignment(vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    tx_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    hc_fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")

    # Заголовки
    headers = [
        ("№", 5),
        ("Тема", 25),
        ("Название материала", 50),
        ("Тип", 15),
        ("Краткое содержание", 70),
        ("Источник / Автор", 25),
        ("Ссылка", 45),
        ("Язык", 8),
        ("Год", 15),
        ("Дата добавления", 15),
        ("Полезность информации", 20),
    ]

    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col_letter}1"
    ws.freeze_panes = "A2"

    # Выпадающий список "Полезность информации" (1-5)
    useful_col_idx = len(headers)
    useful_col_letter = get_column_letter(useful_col_idx)
    useful_validation = DataValidation(
        type="list", formula1='"1,2,3,4,5"', allow_blank=True
    )
    ws.add_data_validation(useful_validation)

    # Данные
    for row_idx, item in enumerate(data, 2):
        row_values = [
            item.get("id", row_idx - 1),
            item.get("topic", ""),
            item.get("title", ""),
            item.get("type", ""),
            item.get("summary", ""),
            item.get("source", ""),
            item.get("url", ""),
            item.get("lang", ""),
            item.get("year", ""),
            item.get("added", ""),
            "",
        ]

        for col_idx, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = link_font if col_idx == 7 else data_font
            cell.alignment = (
                center_align if col_idx in (1, 4, 8, 9, 10, useful_col_idx) else wrap_align
            )
            cell.border = thin_border

            if col_idx == 2:
                if value == "Total Experience":
                    cell.fill = tx_fill
                elif value == "Человекоцентричность":
                    cell.fill = hc_fill

        useful_validation.add(f"{useful_col_letter}{row_idx}")
        ws.row_dimensions[row_idx].height = 45

    # Цветовая индикация "Полезность информации": 1 - красный ... 5 - зелёный
    if data:
        useful_range = f"{useful_col_letter}2:{useful_col_letter}{len(data) + 1}"
        useful_colors = {
            "1": "E06666",
            "2": "F6B26B",
            "3": "FFD966",
            "4": "B6D7A8",
            "5": "6AA84F",
        }
        for val, color in useful_colors.items():
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.conditional_formatting.add(
                useful_range,
                CellIsRule(operator="equal", formula=[f'"{val}"'], fill=fill),
            )

    # Лист статистики
    ws2 = wb.create_sheet("Статистика")
    ws2["A1"] = "Метрика"
    ws2["B1"] = "Значение"
    for c in ["A1", "B1"]:
        ws2[c].font = header_font
        ws2[c].fill = header_fill
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 20

    stats = [
        ("Всего записей", "=COUNTA('База знаний'!A2:A10000)"),
        ("Total Experience", '=COUNTIF(\'База знаний\'!B2:B10000,"Total Experience")'),
        ("Человекоцентричность", '=COUNTIF(\'База знаний\'!B2:B10000,"Человекоцентричность")'),
        ("Статей", '=COUNTIF(\'База знаний\'!D2:D10000,"Статья")'),
        ("Отчётов", '=COUNTIF(\'База знаний\'!D2:D10000,"Отчёт")'),
        ("На русском", '=COUNTIF(\'База знаний\'!H2:H10000,"RU")'),
        ("На английском", '=COUNTIF(\'База знаний\'!H2:H10000,"EN")'),
        ("Последнее обновление", datetime.now().strftime("%Y-%m-%d")),
    ]

    for i, (label, val) in enumerate(stats, 2):
        ws2.cell(row=i, column=1, value=label).font = Font(name="Arial", size=11)
        ws2.cell(row=i, column=2, value=val).font = Font(
            name="Arial", size=11, bold=True
        )

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"XLSX создан: {OUTPUT_FILE} ({len(data)} записей)")


if __name__ == "__main__":
    data = load_data()
    build_xlsx(data)
